from fastapi import FastAPI, UploadFile, File, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates

import os
import shutil
import zipfile

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
import xlrd


app = FastAPI()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
os.makedirs(OUTPUT_DIR, exist_ok=True)

cam_sites = [
    "CAM01", "CAM02", "CAM03", "CAM04", "CAM05", "CAM06",
    "CAM36", "CAM37", "CAM38", "CAM48", "CAM49"
]

# Keep your full ignored codes list here (your repo already has it)
ignored_codes = [
    "CASHOUMACHLAV", "CASNAPBOROVA200M", "CASNAPBOROVA250M", "CASNAPBORREC250M",
    "CASNAPCAR140M", "CASNAPCARCIR160M", "CASNAPOVACIR2M", "CASNAPRECTCIR20M",
    "CASNAPRECTCIR25M", "CASNAPRONCIR140M", "CASNAPRONCIR160M", "CASTABCUIS",
    "DEKBARQALUM1025L", "DEKBARQALUM5069L", "FIMBARGL080", "GEATBARGALU3CP",
    "GEATBARGALUCAK", "GEATBARGALUGM", "GEATBARGALUMM", "GEATBARGALUPM",
    "GEATBOLALUM87", "LBP03BARALUMCKE", "LBP03BARALUPOGM", "LBP04BAR1POC",
    "LBP04BARALUP2", "LBP04BARPIZMD", "LBP04PLASPEGRI", "LBP05BARALUPOPM",
    "LBPBARCAR50CCCV", "LBPBARQALUM504", "LBPLT03BARQALCK", "LBPPALTALUMRECT",
    "UNVBOWSAL69OZ", "UNVECOTRKF1500", "UNVECOTRKF500", "DIVBARPETL101",
    "FIMBARGL332", "LBPBARCAR900GRCV", "LBPBARQPET850L", "TERBARQPET1000",
    "TERBARQPET2000", "UNVBOWSAL43OZ", "GEATCVBARGALUGM", "GEATCVBARGALUMM",
    "ECOPOCHFRITBL", "ECOPOCHMBL155", "ECOPOCHMBL255", "ECOPOCHMBL355",
    "SITPOIGAT120KF", "SITPOIGAT15", "SITPOIGAT20", "DIVBOUGANN00",
    "DIVBOUGANN02", "DIVBOUGANN03", "DIVBOUGANN04", "DIVBOUGANN05",
    "DIVBOUGANN06", "DIVBOUGANN07", "DIVBOUGANN08", "DIVBOUGANN09",
    "DIVBOUTENIGH", "HYP50CVPETTRA35CL", "INN50COUVPET", "VAR50COVPET25",
    "LBPCOUTBOIS", "LBPCUILLBOIS", "LBPFOURBOIS", "DIVCOUTPLAS100",
    "LBP20COUTPLASNO", "WAFCOUPLNOI", "DIVCUISOPLAS100", "DIVPALEGLACOL",
    "GOLTCUIAGLAGLD", "INDCUIACAFTR", "LBP20CUILPLSTRG", "LBP20CUILPLSTRO",
    "LBPAGITCUILCAFTR", "LBPCUIACAF", "LBPCUILASOUP100", "WAFCUILUXTRP",
    "WAFCUISOPNO", "LBP50COUTTRANSP", "LBPCOUTPLAST100", "DIVFOUPLAS100",
    "LBPFOUCPLAST100", "LBPFOUCPLASTBER", "WAFFOURLUXTRP", "LBPSURBLSJETB",
    "HYGTOQMSCHEF", "ALPGANVINYMEDM", "BPGANSPCOT", "INNGANHDPEL",
    "LBPCOUPBOIS12", "LBPPICKBUTTE12", "LBPPICKEVANT", "CASHOUSMICOND",
    "LBPPOCHBAG", "LBPPOCHBAGIMP", "CASSETAPTIS3P", "EMBTAGITAT",
    "ALPBROCHETT20", "ALP100CUREDENTSB", "ALP500CUREDENTSB", "LBPLBCD500P",
    "DIVPAIFELX100", "GOLPAISIMP", "GOLSPAIJUS", "PAP20DEN075",
    "PAP250DEN075", "PAP250DEN085", "MINVERPHYR90P", "WAFVERPYCRI250",
    "WAFVERPYCRI45", "LBP25TASECAPPUC", "VAR50GOBCART12", "VAR50GOBCART18",
    "VAR50GOBCARTFR25", "VAR50GOBCARTFR35", "VAR50GOBCARTKF25", "VAR50GOBCARTVR22",
    "INN50GOBPET10OZ", "INN50GOBPET12OZ", "INN50GOBPET14OZ", "INN50GOBPET16OZ",
    "ADMPAPHYGIMP4", "LBPPAPHGY4CH", "MBAPAPHYGCOM", "MBAPAPYHG300G",
    "EMESSMAI150F", "GEATESSTOUT400", "GEATESSTOUT600", "LBPESSTTLILATTEX",
    "LBPESSTTSCOM400", "LBPPAPESSTCHAR2", "MBAESSTTJAB400", "MBAESSTTJAB750",
    "ADMPAPSERVREST", "GEATPAPSERV3030", "MBASERV3030PP", "MBASERVROURA",
    "ECORPAPCUI500F", "ECOSOPLAIMP32", "TTALUMLAF8M", "TTALUMSTR08M",
    "DEKALUMFOIDA", "TTALUMLAF100M", "TTALUMSTR100M", "INNPAPALU400",
    "TTALUMSTR12M", "TTALUMSTR16M", "TTALUMLAF430", "TTALUMLAF70M",
    "LBPLOALU12FI12G", "INNETIRBLAF08M", "TTETIRBLAF08M", "INNETIRBLAF100M",
    "LBPLOCUI6FI8G", "TTETIRBLAF12M", "TTETIRBLAF16M", "HYPFILETIR200M",
    "INNETIRBV300M", "ECOPAPCUI5MLF", "ECOPAPCUI6M", "BAPSASCONGSB1L",
    "BAPSASCONGSB2L", "HYPSACCUISFOR", "BAPSACPB100LB", "BAPSACPB50L",
    "BAPSACPB70", "BAPSACPB70B", "BAPSACPB70N", "BAPSACPBHD30L"
]

# faster membership test
IGNORED_CODES_SET = set(ignored_codes)

# output formatting
PRODUCTS_PER_COLUMN = 40


def _norm(v):
    if v is None:
        return ""
    return str(v).strip().upper()


def _to_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def _headers_from_row(values):
    headers = {}
    for i, val in enumerate(values, start=1):
        name = str(val).strip() if val is not None else ""
        if name:
            headers[name] = i
    return headers


def _extract_rows_xlsx(path: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("The uploaded Excel file has no active sheet.")

    # pandas header=1 => header is Excel row 2
    header_row_idx = 2
    header_values = [c.value for c in ws[header_row_idx]]
    headers = _headers_from_row(header_values)

    def iter_data_rows():
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            yield row

    return headers, iter_data_rows()


def _extract_rows_xls(path: str):
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)

    # row 2 in Excel => 0-based index 1
    header_row_idx0 = 1
    header_values = sheet.row_values(header_row_idx0)
    headers = _headers_from_row(header_values)

    def iter_data_rows():
        for r in range(header_row_idx0 + 1, sheet.nrows):
            yield sheet.row_values(r)

    return headers, iter_data_rows()


def _read_stock_time(path: str, ext: str) -> str:
    """Read cell F1 (row 1, column 6) from the first sheet."""
    try:
        if ext == ".xlsx":
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            if ws is None:
                return ""
            val = ws.cell(row=1, column=6).value
            return str(val).strip() if val is not None else ""
        elif ext == ".xls":
            book = xlrd.open_workbook(path)
            sheet = book.sheet_by_index(0)
            if sheet.nrows < 1 or sheet.ncols < 6:
                return ""
            val = sheet.cell_value(0, 5)
            return str(val).strip() if val else ""
    except Exception:
        return ""
    return ""


def write_staggered_excel(out_file: str, cam: str, rows: list, stock_time: str = "") -> None:
    """
    Colab-like output:
    - 40 items per column
    - blank separator column between data columns
    - A1 = Site: {cam} | Stock time: {stock_time}
    - row 2 headers = "SFX", "Article", blank, "SFX", "Article", blank, ...
    - data starts row 3: sfx_qty in SFX col, designation in Article col
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    if stock_time:
        ws["A1"] = f"Site: {cam} | Stock time: {stock_time}"
    else:
        ws["A1"] = f"Site: {cam}"

    if not rows:
        ws.cell(row=2, column=1, value="SFX")
        ws.cell(row=2, column=2, value="Article")
        wb.save(out_file)
        return

    num_products = len(rows)
    num_data_columns = (num_products + PRODUCTS_PER_COLUMN - 1) // PRODUCTS_PER_COLUMN

    # headers: each group uses 3 cols (SFX, Article, blank separator)
    for i in range(num_data_columns):
        sfx_col = 1 + i * 3
        article_col = 2 + i * 3
        ws.cell(row=2, column=sfx_col, value="SFX")
        ws.cell(row=2, column=article_col, value="Article")

    # data
    for i in range(num_data_columns):
        start = i * PRODUCTS_PER_COLUMN
        end = min((i + 1) * PRODUCTS_PER_COLUMN, num_products)
        chunk = rows[start:end]

        sfx_col = 1 + i * 3
        article_col = 2 + i * 3
        for r, (sfx_qty, designation) in enumerate(chunk, start=3):
            ws.cell(row=r, column=sfx_col, value=sfx_qty if sfx_qty != "" else None)
            ws.cell(row=r, column=article_col, value=designation)

    wb.save(out_file)


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload/")
async def upload(file: UploadFile = File(...)):
    filename = file.filename or ""
    ext = os.path.splitext(filename.lower())[1]
    temp_path = os.path.join(BASE_DIR, f"temp_{filename}")

    with open(temp_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        # Load rows depending on extension
        try:
            if ext == ".xlsx":
                headers, rows_iter = _extract_rows_xlsx(temp_path)
            elif ext == ".xls":
                headers, rows_iter = _extract_rows_xls(temp_path)
            else:
                return JSONResponse(
                    status_code=400,
                    content={"error": "Unsupported file type. Upload .xls or .xlsx"}
                )
        except (InvalidFileException, zipfile.BadZipFile, xlrd.biffh.XLRDError) as e:
            return JSONResponse(status_code=400, content={"error": "Invalid Excel file.", "details": str(e)})

        stock_time = _read_stock_time(temp_path, ext)

        required = ["Site", "Code", "Désignation Article"]
        missing_cols = [c for c in required if c not in headers]
        if missing_cols:
            return JSONResponse(
                status_code=400,
                content={
                    "error": "Missing required columns in Excel header row (row 2).",
                    "missing": missing_cols,
                    "found": list(headers.keys())
                }
            )

        site_col = headers["Site"]
        code_col = headers["Code"]
        des_col = headers["Désignation Article"]
        stock_col = headers.get("Stock Disponible")  # optional

        def get_cell(row, col_1based):
            i = col_1based - 1
            return row[i] if 0 <= i < len(row) else None

        sfx_products = {}  # (code, des) -> sfx_qty
        cam_products_map = {cam: set() for cam in cam_sites}

        for row in rows_iter:
            site = _norm(get_cell(row, site_col))
            if not site:
                continue

            code = _norm(get_cell(row, code_col))
            des = _norm(get_cell(row, des_col))
            if not code or not des:
                continue

            if site == "SFX":
                if stock_col is not None:
                    stock_val = _to_float(get_cell(row, stock_col))
                    if stock_val is None or stock_val < 4:
                        continue
                    sfx_qty = stock_val
                else:
                    sfx_qty = ""  # blank when Stock Disponible column is absent
                sfx_products[(code, des)] = sfx_qty
            elif site in cam_products_map:
                cam_products_map[site].add((code, des))

        results = []

        sfx_set = set(sfx_products.keys())

        for cam in cam_sites:
            missing = sfx_set - cam_products_map[cam]
            missing = {p for p in missing if p[0] not in IGNORED_CODES_SET}
            # p[1] is already uppercased by _norm, so startswith("PALETTES") is sufficient
            missing = {p for p in missing if not p[1].startswith("PALETTES")}

            rows = sorted([(sfx_products[p], p[1]) for p in missing], key=lambda item: item[1])

            out_file = os.path.join(OUTPUT_DIR, f"{cam}.xlsx")
            write_staggered_excel(out_file, cam, rows, stock_time)

            results.append({"cam": cam, "count": len(rows), "file": f"/download/{cam}.xlsx"})

        zip_path = os.path.join(OUTPUT_DIR, "all_cams.zip")
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
            for cam in cam_sites:
                z.write(os.path.join(OUTPUT_DIR, f"{cam}.xlsx"), f"{cam}.xlsx")

        return {"results": results, "zip": "/download/all"}

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": "Server error", "details": str(e)})

    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


@app.get("/download/{file}")
def download(file: str):
    if file == "all":
        zip_path = os.path.join(OUTPUT_DIR, "all_cams.zip")
        if not os.path.isfile(zip_path):
            return JSONResponse(status_code=404, content={"error": "No processed ZIP yet. Upload a file first."})
        return FileResponse(zip_path, filename="all_cams.zip")

    path = os.path.join(OUTPUT_DIR, file)
    if not os.path.isfile(path):
        return JSONResponse(status_code=404, content={"error": "File not found. Upload a file first."})
    return FileResponse(path, filename=file)
